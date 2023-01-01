import datetime, openpyxl, copy
from openpyxl import load_workbook

#start and end dates
def DateRange(sheet):

    start_date = sheet["A5"].value
    end_date = None

    for row in sheet.iter_rows():
        try:
            if row[0].value > start_date:
                end_date = row[0].value
        except:
            pass
    end_date += datetime.timedelta(days = 10)

    dateList = []
    date = start_date

    while date != end_date + datetime.timedelta(days = 1):
        dateList.append(date)
        date += datetime.timedelta(days = 1)

    return start_date, end_date, dateList


#adding dates
def DateAdd(sheet, date_row, min_column):
    global DateList
    dateList2 = copy.deepcopy(DateList)

    #adding dates to columns
    for column in sheet.iter_cols(min_col = min_column, max_col= len(DateList)+8, min_row = date_row, max_row = date_row):
        for cell in column:
            cell.value = dateList2.pop(0)



#project ammounts to the correct columns
def Project(sheet, max_row, start_row, date_row):
    global DateList, weekday

    for r in range(start_row, max_row):
        if weekday == True:
            date = dateAdjust(r)
        else:
            date = PD.cell(row = r, column = 1).value
        ammount = sheet.cell(row = r, column = 9).value

        for c in range(1, len(DateList)+1):
            cross_checker = sheet.cell(row = date_row, column = c).value
            if cross_checker == date:
                try:
                    cell = sheet.cell(row = r, column = c)
                    cell.value = ammount
                except:
                    pass


def dateAdjust(r):
    global CIH
    
    thedate = CIH.cell(row = r, column = 1).value

    holidays = []

    wb = load_workbook(filename = "Data.xlsx" )
    hols = wb['Holidays']
    for row in hols.iter_rows():
        holidays.append(row[1].value)
    wb.close()


    thedate = NonWorkingDay(thedate, holidays)

    bank = CIH.cell(row = r, column = 4).value
    if bank == "Nations Trust Bank Plc (Colombo)":
        thedate = thedate
    else:
        thedate += datetime.timedelta(days = 1)
    
    thedate = NonWorkingDay(thedate, holidays)


    return thedate

def NonWorkingDay(thedate, holidays):
    try:
        while (thedate in holidays) or (thedate.weekday() == 6) or (thedate.weekday() == 5):
            thedate += datetime.timedelta(days = 1)
    except:
        pass
    return thedate



#sums all columns
def sum(max_row, sheet, first_row):
    last_row = max_row - 1
    for column in sheet.iter_cols(min_col= 8, min_row= max_row, max_row = max_row):
        for cell in column:
            argument1 = cell.column_letter + str(first_row)
            argument2 = cell.column_letter + str(last_row)
            cell.value = '=sum({0}:{1})'.format(argument1, argument2)

def dateFormat(sheet, firstrow):
    for row in sheet.iter_rows(min_col=1, max_col=1, min_row=firstrow, max_row=sheet.max_row-1):
        for cell in row:
            cell.number_format = 'Short Date'


def balances(sheet, start_column, start_row, end_row, cb_row):
    #for opening balances
    for column in sheet.iter_cols(min_col= start_column+1, min_row = 4, max_row = 4):
        for cell in column:
            letter = chr(ord(cell.column_letter) - 1)
            cell.value = letter + "23"

    #for closing balance 1
    for column in sheet.iter_cols(min_col= start_column, min_row= cb_row, max_row = cb_row):
        for cell in column:
            argument1 = cell.column_letter + str(start_row)
            argument2 = cell.column_letter + str(end_row)
            cell.value = '=sum({0}:{1})'.format(argument1, argument2)
    
    #for cash in hand
    for column in sheet.iter_cols(min_col= start_column, min_row= cb_row+3, max_row = cb_row+3):
        for cell in column:
            cell.value = "=$B$2"

    #for closing balance 2
    for column in sheet.iter_cols(min_col = start_column, min_row= cb_row+5, max_row = cb_row+5):
        for cell in column:
            cell.value = cell.column_letter + "23 + " + cell.column_letter + "25"

def constants(sheet):
    for column in sheet.iter_cols():
        if column[2].value.date == 5:
            column[8].value = "B8"
            column[9].value = "B9"
            column[10].value = "B10"
            column[11].value = "B11"
            column[12].value = "B12"
            column[14].value = "B14"

        


excel = load_workbook(filename = "CIH.xlsx" )
CIH = excel.active

weekday = True

start_date, end_date, DateList = DateRange(CIH)
DateAdd(CIH, 2, 10)
Project(CIH, CIH.max_row, 5, 3)
sum(CIH.max_row, CIH, 5)

excel.save(filename = "Report1.xlsx")




excel = load_workbook(filename = "PD.xlsx")
PD = excel.active

weekday = False

dateFormat(PD, 7)
DateAdd(PD, 4, 10)
Project(PD, PD.max_row, 7, 5)
sum(PD.max_row, PD, 7)

excel.save(filename= "Report2.xlsx")




excel = load_workbook(filename = "Data.xlsx")
C_Flow = excel["C Flow"]

DateAdd(C_Flow, 2, 3)
balances(C_Flow, 3, 4, 21, 23)
constants(C_Flow)

excel.save(filename= "Report3.xlsx")